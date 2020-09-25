#!/usr/bin/python3
################################################################################
#
# Description: This script is part of domainBlender and is used to merge the
#              data of typofinder, squatcobbler and urlcrazy.
#
# File dependencies: The following files are required for the script to run
#                    optimally. The name of the files can be adjusted in the
#                    top section of this script.
#                  - <domainlist_file>: list of domains to be checked. The
#                                       format of the file should be of one
#                                       domain per line.
#                  - <alltld_file>: file containing a list of all the valids
#                                   TLDs. Default is effective_tld_names.dat
#                                   from typofinder.
#                  - <whitelist_file>: file of domains to be removed from the
#                                      results. For domain and related
#                                      subdomains, use the format "*.<domain>".
#                  - <category_file>: csv file containing the categorization of
#                                     a domain.
#                  - <results_location_per_tool>: file(s) containing the
#                                                 results of the tools. Can be
#                                                 either a filename.
#
# File created: The following file is created and/or replaced by the script
#               when ran.
#               - <results_target>: Excel spreadsheet (.xlsx) resulting of the
#                                   combination and correlation of the
#                                   information provided.
#
# Comments: - To add support to another tool, add the name of the tool to the
#             variable "results_location_per_tool". Also, create a function
#             with the same name to parse the data (see already made parser of
#             typofinder, urlcrazy and squatcobbler for examples).
#           - Although the result files in results_location_per_tool can be a
#             folder, it hasn't been tested and may need adjustments.
#
# Created by: Jean-Francois Boucher
# Date of creation: 10 Jul 2020
#
# Modified by: Jean-Francois Boucher
# Date of modification: 17 Jul 2020
#
################################################################################


##### Parameters to adjust #####

domainlist_file = "domains"
results_target = "Result_typo_domains.xlsx"
alltld_file = "./typofinder/TypoMagic/datasources/effective_tld_names.dat"
whitelist_file = "whitelist.txt"
category_file = "categories.csv"

# Format of "results_location_per_tool":
#    For a file: {"toolname1": "results_filename"}
#    For a folder: {"toolname2": ["results_foldername", "filename_pattern"]}
# Example: {"tool1": "/tmp/tool1results.json", "tool2": ["./tool2results/", "*.csv"]}
results_location_per_tool= {"typofinder": "./typofinder.json",
                            "urlcrazy": "./urlcrazy_results.csv",
                            "squatcobbler": "./squatcobbler_results.json"}

##### Parameters to adjust end #####


##### Import start #####

import ast
#import csv
import glob
import json
import os
import sys

from openpyxl import Workbook

##### Import end #####

##### global variables start #####

# List of required fields. Also determine the order of the results
# Note: do NOT use the name "extra_fields" as it is used internally for extra
#       values of the other fields
required_fields = ["typo_domain",
                   "main_domain",
                   "TLD",
                   "original_domain",
                   "A_record",
                   "NS_record",
                   "MX_record",
                   "country",
                   "country_code",
                   "source_tool",
                   "category",
                   "registrar",
                   "creation",
                   "update",
                   "expiration"]

toollist= []

##### global variables end #####


##### verify_dependencies start #####

def verify_dependencies():
    '''verify_dependencies() -> dict

    Function to verify that all files required are presents. It also verify for
    the optional files (i.e. whitelist_file).'''

    present={"domainlist": False,
             "alltld_file": False,
             "whitelist_file": False,
             "category_file": False}

    filemapping={"domainlist": domainlist_file,
                 "alltld_file": alltld_file,
                 "whitelist_file": whitelist_file,
                 "category_file": category_file}

    # Verify the result file can be created
    # Verify results_target is just a filename or a path
    if results_target.find("/") == -1:
        target="."
    else:
        target=results_target.rsplit("/",1)[0]

    # Verify the path exists
    if os.path.exists(target):
        # Verify the target is a directory
        if os.path.isdir(target):
            # Verify the permissions permits to create/read/write the file
            try:
                f=open(results_target, "w")
                f.write("test")
                f.close()
                f=open(results_target)
                test_string=f.read()
                f.close()
                if test_string != "test":
                    print("ERROR: Cannot read and/or write file " + results_target + ". QUITTING!")
                    sys.exit()
                else:
                    os.remove(results_target)
            except IOError:
                print("ERROR: Issues accessing file " + results_target + ". Verify permissions. QUITTING!")
                sys.exit()

    # Verify the presence of the input files
    for target in [*filemapping.keys()]:
        # Verify the path exists
        if os.path.exists(filemapping[target]):
            # Verify it is a file
            if os.path.isfile(filemapping[target]):
                    present[target] = True
    return present

##### verify_dependencies end #####


##### read_txtfile start #####

def read_txtfile(filename):
    '''read_txtfile(filename) -> list

    Function to read a text file and return the results. Doesn't parse the
    formatting of the lines (i.e. csv) but remove the carriage returns and
    empty lines.'''

    content = []

    # Read the file
    f = open(filename)
    tmpcontent = f.readlines()
    f.close()

    # Loop to process each lines
    for lines in tmpcontent:
        # Skip empty lines
        if not lines in ["\n", "\r\n", "\r"]:
            # Skip commented lines
            if lines.find("/") != 0:
                content.append(lines.strip("\n").strip("\r"))
    return content

##### read_txtfile end #####


##### list_files start #####

def list_files(source):
    '''list_fil(list/str) -> list

    Function to get the list of result files for a tool.'''

    file_list=[]
    errors=""

    # Verify the source is a str or a list
    if type(source) == str:
        str_source=source
    else:
        str_source=source[0]

    # Verify the path exists
    if not os.path.exists(str_source):
        errors = "Path '" + str_source + "'doesn't exist"

    # Verify the path is a directory
    elif os.path.isdir(str_source):

        # Validate if the source had a "/" or not
        if type(source) == list:
            if source[0].rfind("/") == len(source[0] - 1):
                sep=""
            else:
                sep="/"

            # Get the lists of file matching the directory and pattern
            tmp_file_list=glob.glob(source[0] + sep + source[1])
        else:
            errors = "Path '" + str_source + "'is a folder although no filename nor filename pattern was provided"

    # Verify the source is a str
    elif type(source) == str:
        # Convert the str into a list
        file_list=[source]
    else:
        errors = "Path '" + str_source + "' is a file although a foldername format was provided"

    return file_list, errors

##### list_files end #####


##### get_categories start #####

def get_categories():
    '''get_cateogies() -> dict

    Function to read the category .csv file and parse it into a dict.'''

    all_categories = {}

    # Read the file
    content=read_txtfile(category_file)

    # Loop through all the lines for parsing fields
    for line in content:
        # Parse the csv lines
        domain, category, source = line.split(",")

        #Verify if the key needs to be created and/or the value added
        if not domain in all_categories.keys():
            all_categories[domain]= {}
        if not "category" in all_categories[domain].keys():
            all_categories[domain]["category"]=[]
        if not category in all_categories[domain]["category"]:
            all_categories[domain]["category"].append(category)
        if not "source" in all_categories[domain].keys():
            all_categories[domain]["source"]=[]
        if not source in all_categories[domain]["source"]:
            all_categories[domain]["source"].append(source)

    return all_categories

##### get_categories end #####


##### typofinder start #####

def typofinder(filename, all_categories, alltld_list):
    '''typofinder(filename, all_categories, alltld_list) -> dict

    Function to parse the results of typofinder to standardize the fields.'''

    field_mapping = {"typo_domain": "domain",
                     "main_domain": "core_domain",
                     "TLD": "tld",
                     "A_record": "IPv4Addresses",
                     "NS_record": "nameservers",
                     # Note: To have MX_record as IPv4, change to "aMXIP4"
                     "MX_record": "aMX",
                     "country": "whois",
                     "country_code": "whois",
                     "source_tool": "",
                     "category": "category",
                     "registrar": "whois",
                     "creation": "whois",
                     "update": "whois",
                     "expiration": "whois",
                     "original_domain": "original_domain"}

    # Create a dictionary to parse the whois information
    whois_mapping = {}
    whois_mapping["registrant"] = {"country": " country",
                                    "country_code": "country"}
    whois_mapping["registrar"] = {"registrar": "name"}
    whois_mapping["date"] = {"creation": "created",
                                    "update": "updated",
                                    "expiration": "expires"}

    tmpresults = {}

    # import json file's data
    f = open(filename)
    rawresults = json.load(f)
    f.close()

    # extract the keys of the dict
    domainlist = [*rawresults.keys()]

    # Loop to parse each entries of the file to match the standardized format
    for domain in domainlist:

        # Verify the domain is not already in tmpresults
        if not domain in tmpresults:
            tmpresults[domain] = {}

            # Split the domain name
            subdomain, core_domain, tld = explode_domain(domain, alltld_list)

            # Loop to process all the required fields
            for fields in field_mapping.keys():

                if fields == "typo_domain":
                    tmpresults[domain]["typo_domain"] = domain

                elif fields == "main_domain":
                    tmpresults[domain]["main_domain"] = core_domain + "." + tld

                elif fields == "TLD":
                    tmpresults[domain]["TLD"] = tld

                elif field_mapping[fields] == "category":
                    # Verify if the domain is in the category.
                    # Example: if domain is "www.forces.gc.ca",
                    #          core_domain=forces,
                    #          tld=gc.ca,
                    #          check in order "www.forces.gc.ca", then
                    #          "forces.gc.ca", then "gc.ca"
                    if domain in all_categories.keys():
                        tmpresults[domain][fields] = all_categories[domain]["category"].copy()
                    elif core_domain + "." + tld in all_categories.keys():
                        tmpresults[domain][fields] = all_categories[core_domain + "." + tld]["category"].copy()
                    elif tld in all_categories.keys():
                        tmpresults[domain][fields] = all_categories[tld]["category"].copy()
                    else:
                        tmpresults[domain][fields] = ""

                # Uncomment this section to have the MX_record as IPv4 addresses
                # To extract the IPs of the MX records instead of names
                #elif fields == "MX_record":
                #    if "aMXIPv4" in rawresults[domain].keys():
                #        if len(rawresults[domain]["aMXIPv4"]) != 0:
                #            if type(rawresults[domain]["aMXIPv4"]) == dict:
                #                mx = []
                #                mx_names=[*rawresults[domain]["aMXIPv4"].keys()]
                #                for mx_name in mx_names:
                #                    for item in rawresults[domain]["aMXIPv4"][mx_name]:
                #                        if not item in mx:
                #                            mx.append(item)
                #                tmpresults[domain]["MX_record"] = mx
                #        else:
                #            tmpresults[domain]["MX_record"]=""
                #    else:
                #        tmpresults[domain]["MX_record"] = ""

                elif fields == "source_tool":
                     tmpresults[domain][fields] = "typofinder"

                elif field_mapping[fields] == "whois":
                    tmpresults[domain][fields] = ""

                else:
                    tool_field = field_mapping[fields]

                    # Verify the field is required/mapped
                    if tool_field != "":

                        # Verify the field has a value
                        if tool_field in rawresults[domain].keys():
                           tmpresults[domain][fields] = rawresults[domain][tool_field]
                        else:
                           tmpresults[domain][fields] = ""
                    else:
                        tmpresults[domain][fields] = ""

            # Verify if rawresults contain a whois
            if "whois" in rawresults[domain].keys():

                # Loop to process desired keys of the whois
                for category in whois_mapping.keys():
                    # Verify category exists as of whois key
                    if category in rawresults[domain]["whois"].keys():

                        # Loop to process desired fields of the whois
                        for field in whois_mapping[category].keys():
                            # Verify the field exist in the whois
                            if field in rawresults[domain]["whois"][category].keys():
                                # Verify the field has a value
                                if len(rawresults[domain]["whois"][category][field]) != 0:
                                    tmpresults[domain][whois_mapping[category][field]] = rawresults[domain]["whois"][category][field]

    return tmpresults

##### typofinder end #####


###### urlcrazy start #####

def urlcrazy(filename, all_categories, alltld_list):
    '''urlcrazy(filename, all_categories, alltld_list) -> dict

    Function to parse the results of urlcrazy to standardize the fields.'''

    field_mapping = {"typo_domain": "Typo",
                     "main_domain": "main_domain",
                     "TLD": "tld",
                     "A_record": "DNS-A",
                     "NS_record": "DNS-NS",
                     "MX_record": "DNS-MX",
                     "country": "Country",
                     "country_code": "CountryCode",
                     "source_tool": "",
                     "category": "category",
                     "registrar": "",
                     "creation": "",
                     "update": "",
                     "expiration": "",
                     "original_domain": "original_domain"}

    tmpresults = {}

    # read the file
    content=read_txtfile(filename)
    raw_results=[]

    # Verify the lines start with '"' to remove comments.
    for line in range(0, len(content)):
        # Note: assume a line starting with a '"' is a data line
        if content[line].find('"') ==0:
            raw_results.append(ast.literal_eval(content[line]))

    # Extract the first line for the column header. It is also used to
    # determine where the original domain changed in the file since
    # urlcrazy append the results of each original domain to the same file.
    header=raw_results.pop(0)
    newdomain=True

    # Loop to process each lines
    for line in raw_results:

        # Due to the way domainBlender handles urlcrazy results, verify if the
        # line symbolize the beginning of a new set results.
        if line == header:
            newdomain=True

        else:
            # Verify if it's a new original domain queried, extract the name
            # of the original domain to append to the results.
            if newdomain:
                original_domain=line[1]
                newdomain=False

            # Extract the domain name
            domain=line[1]

            values={}

            # Loop to convert the line into a dict
            for field in range(0, len(header)):
                values[header[field]]=line[field]

            # Verify if the field Valid is true (Note Valid=str)
            if values["Valid"] == "true":
                empty = True

                # Loop to process each fields to see if the results have no
                # significant values
                for field in header:
                    if field in ["DNS-A", "Country", "CountryCode", "DNS-NS", "DNS-MX"]:
                        if len(values[field]) != 0:
                            empty=False

                # Verify if the important fields were not empty
                if empty==False:

                    # Verify if the domain needs to be created in tmpresults.
                    # Also, since a domain can be multiple times in the results
                    # but seems to always have the same values (except the type
                    # of typo and original domain), skips the new results.
                    if not domain in tmpresults.keys():
                        tmpresults[domain] = {}

                        # Split the domain name
                        subdomain, core_domain, tld = explode_domain(domain, alltld_list)

                        # Loop to process all the required fields
                        for fields in field_mapping.keys():

                            if fields == "typo_domain":
                                tmpresults[domain]["typo_domain"] = domain

                            elif fields == "main_domain":
                                tmpresults[domain]["main_domain"] = core_domain + "." + tld

                            elif fields == "TLD":
                                tmpresults[domain]["TLD"] = tld

                            elif fields == "original_domain":
                                tmpresults[domain]["original_domain"] = original_domain

                            elif field_mapping[fields] == "category":
                                # Verify if the domain is in the category.
                                # Example: if domain is "www.forces.gc.ca",
                                #          core_domain=forces,
                                #          tld=gc.ca,
                                #          check in order "www.forces.gc.ca", then
                                #          "forces.gc.ca", then "gc.ca"
                                if domain in all_categories.keys():
                                    tmpresults[domain][fields] = all_categories[domain]["category"].copy()
                                elif core_domain + "." + tld in all_categories.keys():
                                    tmpresults[domain][fields] = all_categories[core_domain + "." + tld]["category"].copy()
                                elif tld in all_categories.keys():
                                    tmpresults[domain][fields] = all_categories[tld]["category"].copy()
                                else:
                                    tmpresults[domain][fields] = ""

                            elif fields == "source_tool":
                                tmpresults[domain][fields] = "urlcrazy"

                            else:
                                tool_field = field_mapping[fields]

                                # Verify the field is required/mapped
                                if tool_field != "":
                                    if tool_field in values.keys():

                                        # Verify the field has a value
                                        if len(values[tool_field]) != 0:
                                            # Verify the value doesn't end by
                                            # a "."
                                            if values[tool_field].rfind(".") == len(values[tool_field]) -1:
                                                tmpresults[domain][fields] = values[tool_field].rsplit(".",1)[0]
                                            else:
                                               tmpresults[domain][fields] = values[tool_field]
                                        else:
                                            tmpresults[domain][fields] = ""
                                    else:
                                      tmpresults[domain][fields] = ""
                                else:
                                    tmpresults[domain][fields] = ""

    return tmpresults

###### urlcrazy end #####


###### squatcobbler start #####

def squatcobbler(filename, all_categories, alltld_list):
    '''squatcobbler(filename, all_categories, alltld_list) -> dict

    Function to parse the results of squat cobbler to standardize the fields.'''

    field_mapping = {"typo_domain": "domain",
                     "main_domain": "core_domain",
                     "TLD": "tld",
                     "A_record": "IPaddr",
                     "NS_record": "",
                     "MX_record": "",
                     "country": "",
                     "country_code": "",
                     "source_tool": "",
                     "category": "",
                     "registrar": "Registrar",
                     "creation": "Created",
                     "update": "Updated",
                     "expiration": "",
                     "original_domain": "Original"}

    rawresults = {}
    tmpresults = {}

    # Import json file's data
    f = open(filename)
    content = json.load(f)
    f.close()

    # Reorganize the results in a dict with the typo domain as key
    for lines in content:
        domain = lines["Modified"]
        if not domain in rawresults.keys():
            rawresults[domain]=lines

    # Extract the keys of the dict
    domainlist = [*rawresults.keys()]

    # Parse each entries of the file to match the standardized format
    for domain in domainlist:
        if not domain in tmpresults:
            tmpresults[domain] = {}

            # Split the domain name
            subdomain, core_domain, tld = explode_domain(domain, alltld_list)

            # Loop to process all the required fields
            for fields in field_mapping.keys():

                if fields == "typo_domain":
                    tmpresults[domain]["typo_domain"] = domain

                elif fields == "main_domain":
                    tmpresults[domain]["main_domain"] = core_domain + "." + tld

                elif fields == "TLD":
                    tmpresults[domain]["TLD"] = tld

                elif field_mapping[fields] == "category":
                    # Verify if the domain is in the category.
                    # Example: if domain is "www.forces.gc.ca",
                    #          core_domain=forces,
                    #          tld=gc.ca,
                    #          check in order "www.forces.gc.ca", then
                    #          "forces.gc.ca", then "gc.ca"
                    if domain in all_categories.keys():
                        tmpresults[domain][fields] = all_categories[domain]["category"].copy()
                    elif core_domain + "." + tld in all_categories.keys():
                        tmpresults[domain][fields] = all_categories[core_domain + "." + tld]["category"].copy()
                    elif tld in all_categories.keys():
                        tmpresults[domain][fields] = all_categories[tld]["category"].copy()
                    else:
                        tmpresults[domain][fields] = ""

                elif fields == "source_tool":
                     tmpresults[domain][fields] = "squatcobbler"

                else:
                    tool_field = field_mapping[fields]
                    if tool_field != "":
                        if tool_field in rawresults[domain].keys():
                           tmpresults[domain][fields] = rawresults[domain][tool_field]
                        else:
                           tmpresults[domain][fields] = ""
                    else:
                        tmpresults[domain][fields] = ""

    return tmpresults

###### squatcobbler end #####


##### merge_results start #####

def merge_results(results, tmpresults):
    '''merge_results(results, tmpresults) -> dict

    Function to combine different sets of results into one dict.'''

    # Loop to process all tmpresults
    while len(tmpresults) > 0:

        # Extact the first key of tmpresults (domain)
        result=[*tmpresults.keys()][0]

        # Verify the domain is not already in the results
        if not result in results.keys():
            results[result]=tmpresults.pop(result)
        else:

            # Compare all the required fields of the domain in tmpresults to
            # the one in results
            for field in required_fields:
                # Verify if the field in results is empty
                if len(results[result][field]) == 0:
                    # Verify if the field in tmpresults has a value
                    if len(tmpresults[result][field]) != 0:
                        # Add the value to results
                            results[result][field] = tmpresults[result][field]

                # results field has a value
                else:
                    # Verify if the field in tmpresults has a value
                    if len(tmpresults[result][field]) != 0:
                        # Verify the type to perform proper search to avoid duplicates
                        # Verify for lists
                        if type(tmpresults[result][field]) == list:
                            # Convert the field in result to a list
                            if type(results[result][field]) == str:
                                results[result][field]=[results[result][field]]

                            if type(results[result][field]) == list:
                                for item in tmpresults[result][field]:
                                    if not item in results[result][field]:
                                        results[result][field].append(item)

                        # Verify for strings
                        elif type(tmpresults[result][field]) == str:

                            if type(results[result][field]) == list:
                                if not tmpresults[result][field] in results[result][field]:
                                    results[result][field].append(tmpresults[result][field])

                            elif type(results[result][field]) == str:
                                if tmpresults[result][field] != results[result][field]:
                                    # Convert the field in results to a list
                                    results[result][field]=[results[result][field]]
                                    results[result][field].append(tmpresults[result][field])

            # TODO verify if there is a need for having a list of original
            #      domains associated to the typo (i.e. [read original=typo]
            #      forces.gc.ca=forces.co and forces.ca=forces.co by changing
            #      TLDs)
            garbage=tmpresults.pop(result)

    return results

##### merge_results end #####


##### explode_domain start #####

def explode_domain(domain, alltld_list):
    '''explode_domain(domain, alltld_list) -> str, str, str

    Function to split a domain to obtain the subdomain, the core domain and the
    TLD'''

    core_domain = ""
    tld = ""
    subdomain = ""

    tmp_tldlist = []

    # Loop to test with all the TLDs
    for tmp_tld in alltld_list:

        # Get the location of the TLD pattern in the domain name
        findloc=domain.rfind(tmp_tld)

        # Verify tmp_tld is in the domain
        if findloc != -1:

            # Verify tmp_tld is not the full value of domain
            # (i.e. if domain="gc.ca")
            if findloc != 0:
                domain_length=len(domain)
                tld_length=len(tmp_tld)

                # Verify the string found in the domain is at the end (tld)
                if findloc + tld_length == domain_length:

                    # Verify the domain and tld are separated by "."
                    if domain[domain_length - tld_length -1] == ".":

                        # Verify if the tmp_tld is longer than a previously
                        # found tld to catch long tld such as "gc.ca" vice
                        # "ca")
                        if len(tld) < len(tmp_tld):
                            tld = tmp_tld
                            tmp_domain = domain.rsplit("." + tld, 1)[0]

                            # Verify the domain includes a subdomain
                            findloc=tmp_domain.rfind(".")
                            if findloc != -1:
                                core_domain = tmp_domain.rsplit(".", 1)[1]
                                subdomain = tmp_domain.rsplit("." + core_domain, 1)[0]

                            else:
                                core_domain = tmp_domain
                                subdomain = ""

    return subdomain, core_domain, tld

##### explode_domain end #####


##### rm_orig_whitelist start #####

def rm_orig_whitelist(results, present):
    '''rm_orig_whitelist(results, present) -> dict

    Function to remove original domains from the results. Also remove known
    (whitelisted) domains from the results.'''
    orig_whitelist = []

    # Verify the domainlist file exist
    if present["domainlist"]:

        # Read the domainlist_file
        orig_whitelist=read_txtfile(domainlist_file)

    # Verify the whitelist file exist
    if present["whitelist_file"]:
        # Read the whitelist_file
        content = read_txtfile(whitelist_file)

        # Add the values from the whitelist to the domainlist
        for lines in content:
             # Verify the value doesn't already exist
             if not lines in orig_whitelist:
                 orig_whitelist.append(lines)

    # Verify there is something in orig_whitelist
    if len(orig_whitelist) != 0:

        # Loop to process all the values from the orig_whitelist
        for domain in orig_whitelist:

            # Verify if the wildcard "*." is present at the beginning of the
            # value
            if domain.find("*.")==0:
                exact=False
                pattern=domain.strip("*.")
            else:
                exact=True
                pattern=domain

            # Verify if the pattern is in the results (exact match)
            if pattern in results.keys():

                # Remove the domain from the results
                garbage=results.pop(pattern)

            # Verify if the wildcard was used
            elif exact==False:
                length=len("." + pattern)

                # Extract the domains from the results
                keys=[*results.keys()]

                # Loop to process all the domains in the results
                for key in keys:
                    # Verifyif the pattern exist in the domain name
                    if key.rfind("."+ pattern) != -1:
                        # Verify that the pattern is at the end the domain name
                        if key.rfind("."+ pattern) + length == len(key):
                            garbage=results.pop(key)

    return results

##### rm_orig_whitelist end #####


##### prep_excel_results start #####

def prep_excel_results(results):
    '''prep_excel_results(results) -> dict, dict

    Function to format the results to be exported. To be able to manipulate
    data in excel in a simple way, a cell can only hold one value, not lists.
    To accomodate for this while keeping the extra values, this function
    convert the lists and dicts into str but create a dict of the extra values.
    A list of extra values is also generated so if/when adding the values to a
    spreadsheet in extra columns, the order of the column can be maintained.'''

    extras = {}
    list_values = []

    # Extract a list of all the domains in the results
    all_domains = [*results.keys()]
    # Loop to process all the domains
    for domain in all_domains:

        # Extract the fields related to the domain
        keys=[*results[domain].keys()]
        # Loop to process all the keys
        for key in keys:

            # Verify if the value of the field is a list
            if type(results[domain][key]) == list:

                # Verify the key is source_tool to not create extra column for
                # it
                if key == "source_tool":
                    value = ""
                    # Loop to concatenate all the values in the list into a str
                    for i in range(0, len(results[domain]["source_tool"])):
                        if i !=0:
                            value += ","
                        value += results[domain]["source_tool"][i]
                    results[domain]["source_tool"]=value

                else:
                    list_values = results[domain][key]
                    # Verify if the list is empty
                    if list_values == []:
                        value = ""
                    else:
                        # Extract the first value to save it in the required fields
                        list_values.sort()
                        value = list_values.pop(0)
                    results[domain][key] = value

                    # Verify if other values are in the list
                    if len(list_values) != 0:

                        # Ony keep one value for original domain
                        if not key in ["original_domain"]:
                            # Verify is the key "extra_fields" needs to be
                            # created in the results
                            if not "extra_fields" in results[domain].keys():
                                results[domain]["extra_fields"] = {}

                            # Verify the key doesn't already exist
                            if not key in results[domain]["extra_fields"]:
                                results[domain]["extra_fields"][key] = []

                            # Loop to add all the values to the extra_fields
                            # list of that key
                            for item in list_values:
                                results[domain]["extra_fields"][key].append(item)

                            # Verify if the key already exist in extras (used
                            # in the positioning of the values in the
                            # spreadsheet)
                            if not key in extras.keys():
                                extras[key] = len(list_values)
                            elif extras[key] < len(list_values):
                                extras[key] = len(list_values)

    return results, extras

##### prep_excel_results end #####


##### calc_column start #####

def calc_column(fieldno):
    '''calc_column(fieldno) -> str

    Function to convert the column number into str for excel. ie. column 28 is
    AB'''

    # In a double letters column, get the left letter value
    columnl = fieldno // 26

    # Get the value of the right letter
    columnr = fieldno - (columnl * 26)

    # Verify if the left letter is below "A" in the spreadsheed, i.e. column
    # "B" doesn't have a left letter
    if columnl != 0:
        # Convert the number to the corresponding letter (i.e. 65 = A)
        column = chr(columnl + 64)
    else:
        column = ""

    # Concatenate the letters of the column
    column = column + chr(columnr + 65)
    return column

##### calc_column end #####


##### export_results start #####

def export_results(results, extras):
    '''export_results(results, extras) ->

    Function to export the results to an excel document.'''

    # Create the workbook var
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    row=1

    # Loop to insert the header of each column
    for fieldno in range(0, len(required_fields)):
        column = calc_column(fieldno)
        cell = column + str(row)
        ws[cell] = required_fields[fieldno]

    # Loop to add the header of each column extras
    for extrafields in extras:

        # Loop to add the header of the column as many time as required
        # (i.e. add 5 column if 5 extra value of one type is in the results)
        for i in range(0, extras[extrafields]):
            fieldno += 1

            column=calc_column(fieldno)
            cell = column + str(row)
            ws[cell] = extrafields

    # Loop to process all the results
    while len(results) > 0:
        # Change row
        row+=1

        # Get the domain name list
        domains=[*results.keys()]

        # Sort the list alphabetically and get the first one
        domains.sort()
        domain=domains[0]

        # Loop to process all the required_fields
        for fieldno in range(0, len(required_fields)):

            column=calc_column(fieldno)
            cell = column + str(row)

            fieldname = required_fields[fieldno]

            if fieldname in results[domain].keys():
                ws[cell] = results[domain][fieldname]

        # Verify if there is any extra fields
        if "extra_fields" in results[domain].keys():

            # skip is used for positioning tracking in the columns
            skip=len(required_fields)

            # Loop to process all the extra_fields
            for extrafields in extras:

                # Verify if the field (i.e. NS_record) exists in the
                # "extra_fields"
                if extrafields in results[domain]["extra_fields"].keys():

                    # Loop to process each values of the field
                    for itemno in range(0, len(results[domain]["extra_fields"][extrafields])):

                        # Calculate on which cell the value should be written to
                        curloc=skip+itemno
                        column = calc_column(curloc)
                        cell = column + str(row)
                        ws[cell] = results[domain]["extra_fields"][extrafields][itemno]

                # Make sure of where the point is
                skip=skip+extras[extrafields]

        # Remove the domain from the results dict
        results.pop(domain)

    # Save the file
    wb.save(results_target)


##### main start #####

def main():
    '''def main() ->

    Main function of the script, orchestrating the steps to follow.'''
    # Verify a parameter was provided
    if len(sys.argv)>1:
        # Read input
        toollist=ast.literal_eval(sys.argv[1])

        if type(toollist) == list:

            for tool in toollist:
                if not tool in results_location_per_tool.keys():
                    print("ERROR: This script doesn't support the tool " + tool + ". QUITTING!")
                    sys.exit()

            present=verify_dependencies()
            results = {}
            tmpresults = {}

            all_categories=get_categories()
            alltld_list = read_txtfile(alltld_file)

            for tool in toollist:
                folderlist, errors = list_files(results_location_per_tool[tool])

                if errors !="":
                    print("\nERROR: " + errors + ". Skipping results of " + tool + "\n")
                else:

                    for filename in folderlist:
                        # Call the function named the same as the tool
                        tmpresults = globals()[tool](filename, all_categories, alltld_list)
            
                        results=merge_results(results, tmpresults)

            results = rm_orig_whitelist(results, present)
            results, extras = prep_excel_results(results)

            #print(extras)

            export_results(results, extras)
        else:
            print("ERROR: the parameter provided is not a list, QUITTING")
    else:
        print("ERROR: data_merger.py needs 1 list argument. Use the following command: python3 data_merger.py \"['value1', 'value2']\"")

##### main end #####

# Call main()
main()
